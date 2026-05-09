"""
fetch_pbs_drug_spend.py
━━━━━━━━━━━━━━━━━━━━━━━
Downloads the PBS Expenditure and Prescriptions Excel file and extracts
drug-level spend data — government benefit paid and scripts dispensed per drug.

OUTPUT:
  pbs_drug_spend.csv — one row per drug with:
    drug_name, atc_code, gov_benefit_aud, scripts, cost_per_script_aud,
    report_year, source_url

USAGE:
  python fetch_pbs_drug_spend.py             # latest available year
  python fetch_pbs_drug_spend.py --year 2023 # specific financial year end

SETUP:
  pip install openpyxl requests --break-system-packages
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from pathlib import Path

HERE = Path(__file__).parent

# ── Known Excel URLs by financial year end ────────────────────────────────────
EXCEL_URLS = {
    2024: "https://www.pbs.gov.au/statistics/expenditure-prescriptions/2023-2024/Expenditure-prescriptions-report-tables-2023-24.XLSX",
    2023: "https://www.pbs.gov.au/statistics/expenditure-prescriptions/2022-2023/Expenditure-prescriptions-report-tables-2022-23.XLSX",
    2022: "https://www.pbs.gov.au/statistics/expenditure-prescriptions/2021-2022/Expenditure-prescriptions-report-tables-2021-22.XLSX",
    2021: "https://www.pbs.gov.au/statistics/expenditure-prescriptions/2020-2021/Expenditure-prescriptions-report-tables-2020-21.XLSX",
}

# ── Column name aliases (PBS changes these between years) ─────────────────────
# Maps canonical field → list of possible column header fragments (case-insensitive)
COLUMN_ALIASES = {
    "drug_name":      ["drug name", "medicine name", "generic name", "drug"],
    "atc_code":       ["atc5", "atc code", "atc-5", "atc 5", "atc"],
    "brand_name":     ["brand name", "brand"],
    "gov_benefit":    ["government benefit", "govt benefit", "government cost",
                       "benefit paid", "gov benefit", "government expenditure"],
    "scripts":        ["prescriptions", "scripts", "number of prescriptions",
                       "prescription count", "no. prescriptions"],
    "total_cost":     ["total cost", "total medicine cost", "dispensed cost"],
}

OUTPUT_FIELDS = [
    "drug_name", "brand_name", "atc_code",
    "gov_benefit_aud", "scripts", "cost_per_script_aud",
    "report_year", "source_url",
]


def download_excel(url: str, dest: Path) -> bool:
    """Download file with polite headers. Returns True on success."""
    try:
        import requests
    except ImportError:
        print("Missing requests. Run: pip install requests --break-system-packages")
        sys.exit(1)

    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; PBAC Watch research tool; +https://pbacwatch.vercel.app)",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
    }
    print(f"  Downloading: {url}")
    try:
        r = requests.get(url, headers=headers, timeout=60, stream=True)
        r.raise_for_status()
        with open(dest, "wb") as f:
            for chunk in r.iter_content(chunk_size=65536):
                f.write(chunk)
        size_kb = dest.stat().st_size // 1024
        print(f"  Saved {size_kb} KB → {dest.name}")
        return True
    except Exception as e:
        print(f"  Download failed: {e}")
        return False


def find_column(headers: list[str], field: str) -> int | None:
    """Find column index for a canonical field name. Case-insensitive partial match."""
    aliases = COLUMN_ALIASES.get(field, [field])
    headers_lower = [h.lower().strip() for h in headers]
    for alias in aliases:
        for i, h in enumerate(headers_lower):
            if alias in h:
                return i
    return None


def parse_number(val) -> float | None:
    """Parse a numeric cell value, handling strings with commas/dollar signs."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val != 0 else None
    s = str(val).replace(",", "").replace("$", "").replace(" ", "").strip()
    if not s or s in ("-", "n/a", "na", "—", "*"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def extract_drug_rows(sheet) -> list[dict]:
    """
    Extract drug-level rows from an openpyxl worksheet.
    Scans all rows to find the header row, then extracts data rows.
    """
    from openpyxl import load_workbook

    rows_raw = list(sheet.iter_rows(values_only=True))
    if not rows_raw:
        return []

    # Find header row: the row with the most recognised column names
    header_row_idx = None
    best_score = 0
    for i, row in enumerate(rows_raw[:30]):  # header usually in first 30 rows
        if row is None:
            continue
        cells = [str(c).lower().strip() if c is not None else "" for c in row]
        score = 0
        for field, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                if any(alias in c for c in cells):
                    score += 1
                    break
        if score > best_score:
            best_score = score
            header_row_idx = i

    if header_row_idx is None or best_score < 2:
        return []

    headers = [str(c).strip() if c is not None else "" for c in rows_raw[header_row_idx]]

    # Map canonical fields to column indices
    col = {}
    for field in COLUMN_ALIASES:
        idx = find_column(headers, field)
        if idx is not None:
            col[field] = idx

    if "drug_name" not in col and "atc_code" not in col:
        return []  # can't identify this sheet

    print(f"    Header found at row {header_row_idx+1}. Columns mapped: {list(col.keys())}")

    results = []
    for row in rows_raw[header_row_idx + 1:]:
        if row is None or all(c is None for c in row):
            continue

        def get(field):
            idx = col.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        drug = str(get("drug_name") or "").strip()
        # Strip PBS formulation markers: ^, ^^, *, # at end of name
        drug = re.sub(r'[\^*#]+$', '', drug).strip()
        if not drug or drug.lower() in ("total", "grand total", "subtotal", "drug name",
                                         "all drugs", "all medicines", ""):
            continue
        # Skip rows that look like section headers (all caps, no number in benefit col)
        if drug.isupper() and len(drug) > 30:
            continue

        brand    = str(get("brand_name") or "").strip()
        atc      = str(get("atc_code")   or "").strip()
        benefit  = parse_number(get("gov_benefit"))
        scripts  = parse_number(get("scripts"))

        if benefit is None and scripts is None:
            continue
        if benefit is not None and benefit < 1000:  # noise threshold
            continue

        cost_per = None
        if benefit and scripts and scripts > 0:
            cost_per = round(benefit / scripts, 2)

        results.append({
            "drug_name":         drug,
            "brand_name":        brand,
            "atc_code":          atc,
            "gov_benefit_aud":   int(benefit) if benefit else None,
            "scripts":           int(scripts) if scripts else None,
            "cost_per_script_aud": cost_per,
        })

    return results


def parse_excel(path: Path) -> list[dict]:
    """Parse all sheets in the Excel, return best drug-level table found."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Missing openpyxl. Run: pip install openpyxl --break-system-packages")
        sys.exit(1)

    print(f"  Parsing {path.name} …")
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  Failed to open workbook: {e}")
        return []

    best_rows: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"    Sheet: '{sheet_name}'")
        rows = extract_drug_rows(ws)
        print(f"      → {len(rows)} data rows extracted")
        if len(rows) > len(best_rows):
            best_rows = rows

    wb.close()
    return best_rows


def title_case_drug(name: str) -> str:
    """Convert ALL CAPS drug names to Title Case, preserving known acronyms."""
    if not name:
        return name
    # If already mixed case, leave it
    if not name.isupper():
        return name
    # Known acronyms to keep uppercase
    KEEP_UPPER = {"IV", "PBS", "ATC", "DNA", "RNA", "PEG", "HPV", "BCG", "MMR"}
    words = name.split()
    result = []
    for w in words:
        if w in KEEP_UPPER:
            result.append(w)
        else:
            result.append(w.capitalize())
    return " ".join(result)


def deduplicate_rows(rows: list[dict]) -> list[dict]:
    """
    Aggregate rows with the same drug name (different formulations/brands).
    Sums gov_benefit_aud and scripts; recomputes cost_per_script.
    Keeps the brand name from the highest-spend row.
    """
    agg: dict[str, dict] = {}
    for r in rows:
        key = title_case_drug(r["drug_name"]).lower()
        if key not in agg:
            agg[key] = {
                "drug_name":      title_case_drug(r["drug_name"]),
                "brand_name":     r.get("brand_name", ""),
                "atc_code":       r.get("atc_code", ""),
                "gov_benefit_aud": r.get("gov_benefit_aud") or 0,
                "scripts":         r.get("scripts") or 0,
                "_best_benefit":   r.get("gov_benefit_aud") or 0,
            }
        else:
            entry = agg[key]
            b = r.get("gov_benefit_aud") or 0
            s = r.get("scripts") or 0
            entry["gov_benefit_aud"] = (entry["gov_benefit_aud"] or 0) + b
            entry["scripts"] = (entry["scripts"] or 0) + s
            # Keep brand name from the highest-spend row
            if b > entry["_best_benefit"]:
                entry["_best_benefit"] = b
                entry["brand_name"] = r.get("brand_name", "")
                entry["atc_code"]   = r.get("atc_code", "")

    result = []
    for entry in agg.values():
        benefit = entry["gov_benefit_aud"]
        scripts = entry["scripts"]
        cps = round(benefit / scripts, 2) if scripts and scripts > 0 else None
        if benefit < 1000:
            continue
        result.append({
            "drug_name":           entry["drug_name"],
            "brand_name":          entry["brand_name"],
            "atc_code":            entry["atc_code"],
            "gov_benefit_aud":     int(benefit) if benefit else None,
            "scripts":             int(scripts) if scripts else None,
            "cost_per_script_aud": cps,
        })
    return result


def save_csv(rows: list[dict], path: Path, year: int, url: str):
    """Deduplicate, then write drug spend data to CSV."""
    rows = deduplicate_rows(rows)
    # Sort by government benefit descending
    rows.sort(key=lambda r: r.get("gov_benefit_aud") or 0, reverse=True)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow({
                "drug_name":          r["drug_name"],
                "brand_name":         r.get("brand_name", ""),
                "atc_code":           r.get("atc_code", ""),
                "gov_benefit_aud":    r.get("gov_benefit_aud", ""),
                "scripts":            r.get("scripts", ""),
                "cost_per_script_aud": r.get("cost_per_script_aud", ""),
                "report_year":        year,
                "source_url":         url,
            })
    print(f"  Saved {len(rows)} drugs → {path.name}")


def print_highlights(rows: list[dict]):
    """Print top 10 by spend and top 10 by cost-per-script."""
    print()
    print("  TOP 10 BY GOVERNMENT BENEFIT (2023-24)")
    for i, r in enumerate(rows[:10], 1):
        b = r.get("gov_benefit_aud") or 0
        print(f"    {i:2d}. {r['drug_name']:<35}  ${b/1e6:7.1f}M")

    by_cps = sorted(
        [r for r in rows if r.get("cost_per_script_aud") and r.get("scripts", 0) > 10],
        key=lambda r: r.get("cost_per_script_aud") or 0,
        reverse=True
    )
    print()
    print("  TOP 10 BY COST PER SCRIPT")
    for i, r in enumerate(by_cps[:10], 1):
        cps = r.get("cost_per_script_aud") or 0
        print(f"    {i:2d}. {r['drug_name']:<35}  ${cps:,.0f}/script")


def main():
    parser = argparse.ArgumentParser(description="Fetch PBS drug-level spend data")
    parser.add_argument("--year", type=int, default=max(EXCEL_URLS.keys()),
                        help=f"Financial year end (default: {max(EXCEL_URLS.keys())})")
    parser.add_argument("--keep-xlsx", action="store_true",
                        help="Keep the downloaded Excel file after parsing")
    args = parser.parse_args()

    url = EXCEL_URLS.get(args.year)
    if not url:
        print(f"No URL known for year {args.year}. Add it to EXCEL_URLS.")
        sys.exit(1)

    print("=" * 62)
    print("PBS Drug Spend Fetcher")
    print("=" * 62)
    print(f"  Financial year end : {args.year}")
    print(f"  URL                : {url}")
    print()

    # Download
    xlsx_path = HERE / f"pbs_drug_spend_{args.year}.xlsx"
    if not xlsx_path.exists():
        ok = download_excel(url, xlsx_path)
        if not ok:
            # Try alternate URL pattern (some years use .xlsx not .XLSX)
            alt_url = url.replace(".XLSX", ".xlsx")
            if alt_url != url:
                print(f"  Trying alternate URL: {alt_url}")
                ok = download_excel(alt_url, xlsx_path)
            if not ok:
                print("  Could not download. Check the URL or download manually.")
                sys.exit(1)
        time.sleep(1)  # polite pause
    else:
        print(f"  Using cached {xlsx_path.name}")

    # Parse
    rows = parse_excel(xlsx_path)
    if not rows:
        print()
        print("  No drug-level rows found in any sheet.")
        print("  The Excel structure may have changed. Check the file manually.")
        print(f"  File saved at: {xlsx_path}")
        sys.exit(1)

    # Save
    out_path = HERE / "pbs_drug_spend.csv"
    save_csv(rows, out_path, args.year, url)
    print_highlights(rows)

    # Optionally clean up Excel
    if not args.keep_xlsx:
        xlsx_path.unlink(missing_ok=True)
        print(f"\n  (Excel file removed — use --keep-xlsx to retain it)")

    print()
    print("=" * 62)
    print("  Next: run build_site_data.py to incorporate into the site")
    print("=" * 62)


if __name__ == "__main__":
    main()
